# main.py (Complete with Master Training Database + Hostinger Upload)
from fastapi import FastAPI, HTTPException
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import List, Optional
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
import requests
from io import BytesIO
import os
import shutil
import traceback
import time
from datetime import datetime
from pathlib import Path
from ftplib import FTP

app = FastAPI(title="Excel Export Service")

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Master database configuration
MASTER_DB_PATH = "master/Training_Database_Master.xlsx"
MASTER_DIR = "master"

# Hostinger FTP configuration
HOSTINGER_FTP_HOST = os.environ.get("HOSTINGER_SFTP_HOST", "")
HOSTINGER_FTP_USER = os.environ.get("HOSTINGER_SFTP_USER", "")
HOSTINGER_FTP_PASS = os.environ.get("HOSTINGER_SFTP_PASS", "")
HOSTINGER_REMOTE_DIR = "/files/public_html/uploads/trainees/directory/"
HOSTINGER_PUBLIC_URL = "https://petrosphere.com.ph/uploads/trainees/directory/"

class Trainee(BaseModel):
    id: str
    first_name: str
    last_name: str
    middle_initial: Optional[str] = None
    suffix: Optional[str] = None
    gender: Optional[str] = None
    age: Optional[int] = None
    company_name: Optional[str] = None
    company_position: Optional[str] = None
    company_city: Optional[str] = None
    company_region: Optional[str] = None
    company_industry: Optional[str] = None
    total_workers: Optional[int] = None
    company_email: Optional[str] = None
    email: Optional[str] = None
    phone_number: Optional[str] = None
    company_landline: Optional[str] = None
    picture_2x2_url: Optional[str] = None
    schedule_id: str
    certificate_number: Optional[str] = None

class ExportRequest(BaseModel):
    trainees: List[Trainee]
    courseName: str
    trainingDates: str
    scheduleId: str
    proxyUrl: Optional[str] = None
    eventType: Optional[str] = None
    branch: Optional[str] = None

def upload_master_to_hostinger():
    """Upload the master Excel file to Hostinger via FTP"""
    try:
        if not os.path.exists(MASTER_DB_PATH):
            print("⚠️  Master file doesn't exist, skipping upload")
            return None
        
        if not all([HOSTINGER_FTP_HOST, HOSTINGER_FTP_USER, HOSTINGER_FTP_PASS]):
            print("⚠️  Hostinger FTP credentials not configured")
            return None
        
        print(f"\n{'='*60}")
        print(f"📤 UPLOADING MASTER FILE TO HOSTINGER")
        print(f"{'='*60}")
        print(f"Host: {HOSTINGER_FTP_HOST}")
        print(f"User: {HOSTINGER_FTP_USER}")
        print(f"Remote Directory: {HOSTINGER_REMOTE_DIR}")
        
        # Connect to FTP
        ftp = FTP()
        ftp.connect(HOSTINGER_FTP_HOST, 21, timeout=30)
        ftp.login(HOSTINGER_FTP_USER, HOSTINGER_FTP_PASS)
        
        print(f"✅ Connected to FTP server")
        
        # Navigate to directory
        try:
            ftp.cwd(HOSTINGER_REMOTE_DIR)
            print(f"✅ Changed to remote directory: {HOSTINGER_REMOTE_DIR}")
        except Exception as e:
            print(f"⚠️  Directory doesn't exist, creating: {HOSTINGER_REMOTE_DIR}")
            # Create directory structure if needed
            dirs = HOSTINGER_REMOTE_DIR.strip('/').split('/')
            current_dir = '/'
            for dir_name in dirs:
                current_dir = f"{current_dir}{dir_name}/"
                try:
                    ftp.cwd(current_dir)
                except:
                    ftp.mkd(dir_name)
                    ftp.cwd(current_dir)
        
        # Upload file with timestamp in filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        remote_filename = f"Training_Database_Master_{timestamp}.xlsx"
        
        print(f"📁 Uploading as: {remote_filename}")
        
        with open(MASTER_DB_PATH, 'rb') as file:
            file_size = os.path.getsize(MASTER_DB_PATH)
            print(f"📦 File size: {file_size:,} bytes")
            
            ftp.storbinary(f'STOR {remote_filename}', file)
        
        ftp.quit()
        
        public_url = f"{HOSTINGER_PUBLIC_URL}{remote_filename}"
        
        print(f"✅ Upload successful!")
        print(f"🔗 Public URL: {public_url}")
        print(f"{'='*60}\n")
        
        return public_url
        
    except Exception as e:
        print(f"❌ Error uploading to Hostinger: {str(e)}")
        traceback.print_exc()
        return None

def get_template_filename(trainee_count: int) -> str:
    """Determine which template to use based on trainee count"""
    if trainee_count > 250:
        return 'Directory-300.xlsx'
    elif trainee_count > 200:
        return 'Directory-250.xlsx'
    elif trainee_count > 150:
        return 'Directory-200.xlsx'
    elif trainee_count > 100:
        return 'Directory-150.xlsx'
    elif trainee_count > 50:
        return 'Directory-100.xlsx'
    elif trainee_count > 30:
        return 'Directory-0.xlsx'
    elif trainee_count > 25:
        return 'Directory-1.xlsx'
    elif trainee_count > 20:
        return 'Directory-2.xlsx'
    elif trainee_count > 15:
        return 'Directory-3.xlsx'
    elif trainee_count > 10:
        return 'Directory-4.xlsx'
    else:
        return 'Directory-5.xlsx'

def fetch_image_with_retry(url: str, max_retries: int = 3) -> Optional[bytes]:
    """Fetch image with retry logic"""
    
    # Check if URL is localhost (can't be accessed from Render)
    if 'localhost' in url:
        print(f"   ⚠️  WARNING: URL contains 'localhost' - this won't work from Render!")
        print(f"   💡 Trying to extract actual image URL...")
        
        # Try to extract the actual image URL from the proxy URL
        if '?url=' in url:
            actual_url = url.split('?url=', 1)[1]
            print(f"   ✅ Extracted actual URL: {actual_url[:100]}...")
            url = actual_url
        else:
            print(f"   ❌ Cannot extract actual URL, will fail")
            return None
    
    for attempt in range(max_retries):
        try:
            print(f"   Attempt {attempt + 1}/{max_retries}...")
            
            response = requests.get(
                url, 
                timeout=30,
                headers={
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
                    'Accept': 'image/*,*/*',
                }
            )
            
            if response.status_code == 200:
                return response.content
            else:
                print(f"   HTTP {response.status_code}: {response.reason}")
                
            if attempt < max_retries - 1:
                time.sleep(1)
                
        except requests.exceptions.Timeout:
            print(f"   Timeout on attempt {attempt + 1}")
            if attempt < max_retries - 1:
                time.sleep(1)
        except Exception as e:
            print(f"   Error on attempt {attempt + 1}: {str(e)}")
            if attempt < max_retries - 1:
                time.sleep(1)
    
    return None

def ensure_master_exists():
    """Ensure master training database exists"""
    if not os.path.exists(MASTER_DIR):
        os.makedirs(MASTER_DIR)
        print(f"✅ Created master directory: {MASTER_DIR}")
    
    # If master doesn't exist, create it from template
    if not os.path.exists(MASTER_DB_PATH):
        base_template = "templates/Directory-5.xlsx"
        shutil.copy(base_template, MASTER_DB_PATH)
        print(f"✅ Created new master training database at {MASTER_DB_PATH}")
    
    return MASTER_DB_PATH

def get_next_database_row(db_ws):
    """Find the next available row in Training Database sheet"""
    next_row = 14  # Start from row 14
    
    # Find the last filled row by checking column B (Course Name)
    while db_ws[f'B{next_row}'].value is not None:
        next_row += 1
    
    return next_row

def append_to_master_database(course_name, training_dates, schedule_id, 
                              trainee_count, male_count, female_count, 
                              mode_of_training):
    """Append a new record to the master training database"""
    try:
        master_path = ensure_master_exists()
        
        # Load the master workbook
        master_wb = load_workbook(master_path)
        
        if 'Training Database' not in master_wb.sheetnames:
            print("⚠️  'Training Database' sheet not found in master file")
            master_wb.close()
            return False
        
        db_ws = master_wb['Training Database']
        
        # Find next available row
        next_row = get_next_database_row(db_ws)
        record_number = next_row - 14
        
        print(f"📝 Appending to master database at row {next_row} (Record #{record_number})")
        
        # Write the data
        db_ws[f'A{next_row}'] = record_number
        db_ws[f'B{next_row}'] = course_name
        db_ws[f'C{next_row}'] = training_dates
        db_ws[f'D{next_row}'] = f"#{schedule_id[-4:]}"
        db_ws[f'E{next_row}'] = trainee_count
        db_ws[f'F{next_row}'] = male_count
        db_ws[f'G{next_row}'] = female_count
        db_ws[f'M{next_row}'] = mode_of_training
        db_ws[f'N{next_row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Save the master file
        master_wb.save(master_path)
        master_wb.close()
        
        print(f"✅ Master database updated (Total records: {record_number})")
        
        # Upload to Hostinger
        hostinger_url = upload_master_to_hostinger()
        
        return True
        
    except Exception as e:
        print(f"❌ Error updating master database: {str(e)}")
        traceback.print_exc()
        return False

def copy_master_database_to_export(wb):
    """Copy all records from master database to the export workbook"""
    try:
        if not os.path.exists(MASTER_DB_PATH):
            print("⚠️  Master database doesn't exist yet")
            return False
        
        master_wb = load_workbook(MASTER_DB_PATH)
        
        if 'Training Database' not in master_wb.sheetnames:
            print("⚠️  'Training Database' sheet not found in master")
            master_wb.close()
            return False
        
        if 'Training Database' not in wb.sheetnames:
            print("⚠️  'Training Database' sheet not found in export")
            master_wb.close()
            return False
        
        master_db_ws = master_wb['Training Database']
        export_db_ws = wb['Training Database']
        
        # Copy all data from row 14 onwards
        row_num = 14
        copied_rows = 0
        
        while master_db_ws[f'B{row_num}'].value is not None:
            # Copy each cell from master to export
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'M', 'N']:
                export_db_ws[f'{col}{row_num}'].value = master_db_ws[f'{col}{row_num}'].value
            
            row_num += 1
            copied_rows += 1
        
        master_wb.close()
        print(f"✅ Copied {copied_rows} records from master database to export")
        return True
        
    except Exception as e:
        print(f"⚠️  Error copying master database: {str(e)}")
        return False

@app.get("/")
def read_root() -> dict:
    return {
        "service": "Excel Export Service with Master Database + Hostinger Backup",
        "status": "running",
        "version": "4.0.0",
        "features": [
            "Image proxying", 
            "Hostinger support", 
            "Master database", 
            "Historical records",
            "Automatic Hostinger backup"
        ]
    }

@app.get("/health")
def health_check() -> dict:
    return {"status": "healthy"}

@app.get("/database/stats")
def get_database_stats():
    """Get statistics about the master database"""
    try:
        if not os.path.exists(MASTER_DB_PATH):
            return {
                "status": "success",
                "records": 0, 
                "exists": False,
                "hostinger_configured": bool(HOSTINGER_FTP_HOST and HOSTINGER_FTP_USER and HOSTINGER_FTP_PASS)
            }
        
        wb = load_workbook(MASTER_DB_PATH)
        if 'Training Database' not in wb.sheetnames:
            wb.close()
            return {
                "status": "success",
                "records": 0, 
                "exists": True, 
                "hasSheet": False,
                "hostinger_configured": bool(HOSTINGER_FTP_HOST and HOSTINGER_FTP_USER and HOSTINGER_FTP_PASS)
            }
        
        db_ws = wb['Training Database']
        row_count = get_next_database_row(db_ws) - 14
        wb.close()
        
        return {
            "status": "success",
            "records": row_count,
            "exists": True,
            "hasSheet": True,
            "path": MASTER_DB_PATH,
            "hostinger_configured": bool(HOSTINGER_FTP_HOST and HOSTINGER_FTP_USER and HOSTINGER_FTP_PASS)
        }
    except Exception as e:
        print(f"❌ Error getting database stats: {str(e)}")
        traceback.print_exc()
        return {
            "status": "error",
            "error": str(e)
        }

@app.post("/database/reset")
def reset_master_database():
    """
    Reset the master database to empty state.
    This will backup the current master and create a fresh one.
    """
    try:
        print(f"\n{'='*60}")
        print(f"🔄 DATABASE RESET REQUEST")
        print(f"{'='*60}")
        
        if not os.path.exists(MASTER_DB_PATH):
            print("⚠️  Master database doesn't exist, nothing to reset")
            return {
                "status": "success",
                "message": "Master database doesn't exist, nothing to reset"
            }
        
        # Create backup with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = f"{MASTER_DIR}/Training_Database_Master_BACKUP_{timestamp}.xlsx"
        
        # Backup current master
        shutil.copy(MASTER_DB_PATH, backup_path)
        print(f"✅ Backed up master to: {backup_path}")
        
        # Remove current master
        os.remove(MASTER_DB_PATH)
        print(f"✅ Removed current master")
        
        # Create fresh master from template
        base_template = "templates/Directory-5.xlsx"
        shutil.copy(base_template, MASTER_DB_PATH)
        print(f"✅ Created fresh master database")
        
        # Upload backup to Hostinger
        hostinger_url = upload_master_to_hostinger()
        
        print(f"{'='*60}\n")
        
        return {
            "status": "success",
            "message": "Master database reset successfully",
            "backup_file": backup_path,
            "hostinger_url": hostinger_url,
            "timestamp": timestamp
        }
        
    except Exception as e:
        print(f"❌ Error resetting master database: {str(e)}")
        traceback.print_exc()
        return {
            "status": "error",
            "error": str(e)
        }


@app.post("/database/delete-all-records")
def delete_all_records():
    """
    Delete all records from master database but keep the file structure.
    More dangerous than reset - use with caution!
    """
    try:
        print(f"\n{'='*60}")
        print(f"🗑️  DELETE ALL RECORDS REQUEST")
        print(f"{'='*60}")
        
        if not os.path.exists(MASTER_DB_PATH):
            print("⚠️  Master database doesn't exist")
            return {
                "status": "error",
                "message": "Master database doesn't exist"
            }
        
        # Load the master workbook
        master_wb = load_workbook(MASTER_DB_PATH)
        
        if 'Training Database' not in master_wb.sheetnames:
            master_wb.close()
            return {
                "status": "error",
                "message": "'Training Database' sheet not found"
            }
        
        db_ws = master_wb['Training Database']
        
        # Count existing records
        row_num = 14
        record_count = 0
        while db_ws[f'B{row_num}'].value is not None:
            record_count += 1
            row_num += 1
        
        print(f"Found {record_count} records to delete")
        
        # Delete all data from row 14 onwards
        row_num = 14
        deleted_rows = 0
        while db_ws[f'B{row_num}'].value is not None:
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'M', 'N']:
                db_ws[f'{col}{row_num}'].value = None
            row_num += 1
            deleted_rows += 1
        
        # Save the master file
        master_wb.save(MASTER_DB_PATH)
        master_wb.close()
        
        print(f"✅ Deleted {deleted_rows} records from master database")
        
        # Upload to Hostinger
        hostinger_url = upload_master_to_hostinger()
        
        print(f"{'='*60}\n")
        
        return {
            "status": "success",
            "message": f"Deleted {deleted_rows} records from master database",
            "records_deleted": deleted_rows,
            "hostinger_url": hostinger_url
        }
        
    except Exception as e:
        print(f"❌ Error deleting records: {str(e)}")
        traceback.print_exc()
        return {
            "status": "error",
            "error": str(e)
        }






@app.get("/database/backup")
def backup_master_database():
    """
    Create a backup of the current master database without resetting it.
    """
    try:
        print(f"\n{'='*60}")
        print(f"💾 DATABASE BACKUP REQUEST")
        print(f"{'='*60}")
        
        if not os.path.exists(MASTER_DB_PATH):
            return {
                "status": "error",
                "message": "Master database doesn't exist"
            }
        
        # Create backup with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = f"{MASTER_DIR}/Training_Database_Master_BACKUP_{timestamp}.xlsx"
        
        # Backup current master
        shutil.copy(MASTER_DB_PATH, backup_path)
        file_size = os.path.getsize(backup_path)
        
        print(f"✅ Backed up master to: {backup_path}")
        
        # Also upload to Hostinger
        hostinger_url = upload_master_to_hostinger()
        
        print(f"{'='*60}\n")
        
        return {
            "status": "success",
            "message": "Backup created successfully",
            "backup_file": backup_path,
            "file_size": file_size,
            "hostinger_url": hostinger_url,
            "timestamp": timestamp
        }
        
    except Exception as e:
        print(f"❌ Error creating backup: {str(e)}")
        traceback.print_exc()
        return {
            "status": "error",
            "error": str(e)
        }

@app.post("/export-excel")
def export_excel(request: ExportRequest) -> StreamingResponse:
    try:
        trainees = request.trainees
        course_name = request.courseName
        training_dates = request.trainingDates
        schedule_id = request.scheduleId
        proxy_url = request.proxyUrl
        event_type = request.eventType or 'public'
        branch = request.branch or 'online'
        
        # Determine mode of training
        if event_type.lower() == 'public':
            mode_of_training = f"Public - {branch.title()}"
        elif event_type.lower() == 'in-house':
            mode_of_training = f"In-House - {branch.title()}"
        else:
            mode_of_training = branch.title()

        if not trainees:
            raise HTTPException(status_code=400, detail="At least one trainee is required")

        if not schedule_id or len(schedule_id) < 4:
            raise HTTPException(status_code=400, detail="scheduleId is required")

        trainee_count = len(trainees)
        
        print(f"\n{'='*60}")
        print(f"📊 EXCEL EXPORT REQUEST")
        print(f"{'='*60}")
        print(f"Course: {course_name}")
        print(f"Dates: {training_dates}")
        print(f"Schedule ID: {schedule_id}")
        print(f"Trainee Count: {trainee_count}")
        print(f"Event Type: {event_type}")
        print(f"Branch: {branch}")
        print(f"Mode of Training: {mode_of_training}")
        print(f"Proxy URL: {proxy_url or 'Direct (no proxy)'}")
        print(f"{'='*60}\n")
        
        # Get template
        template_filename = get_template_filename(trainee_count)
        template_path = f"templates/{template_filename}"
        
        print(f"📄 Loading template: {template_filename}")
        
        if not os.path.exists(template_path):
            raise HTTPException(status_code=404, detail=f"Template not found: {template_filename}")
        
        # Load workbook
        wb = load_workbook(template_path)
        ws = wb['Directory of Participants']
        
        # Fill header information
        ws['C10'] = course_name
        ws['C11'] = training_dates
        
        # Count gender
        male_count = sum(1 for t in trainees if t.gender and t.gender.lower() == 'male')
        female_count = sum(1 for t in trainees if t.gender and t.gender.lower() == 'female')
        
        print(f"👥 Gender Distribution: Male={male_count}, Female={female_count}")
        
        # Participant rows start from row 15
        start_row = 15
        images_attempted = 0
        images_successful = 0
        images_failed = 0
        failed_images = []
        
        print(f"\n{'='*60}")
        print(f"📝 PROCESSING TRAINEES")
        print(f"{'='*60}\n")
        
        for i, trainee in enumerate(trainees):
            row_num = start_row + i
            
            print(f"[{i+1}/{trainee_count}] {trainee.first_name} {trainee.last_name}")
            
            # Fill data
            ws[f'A{row_num}'] = i + 1
            ws[f'C{row_num}'] = (trainee.last_name or '').upper()
            ws[f'D{row_num}'] = (trainee.first_name or '').upper()
            
            middle_initial = trainee.middle_initial or ''
            ws[f'E{row_num}'] = middle_initial[0].upper() if middle_initial else ''
            
            ws[f'F{row_num}'] = (trainee.suffix or '').upper()
            ws[f'G{row_num}'] = trainee.gender or ''
            ws[f'H{row_num}'] = trainee.age or ''
            ws[f'I{row_num}'] = trainee.company_name or ''
            ws[f'J{row_num}'] = trainee.company_position or ''
            ws[f'K{row_num}'] = trainee.company_city or ''
            ws[f'L{row_num}'] = trainee.company_region or ''
            ws[f'M{row_num}'] = trainee.company_industry or ''
            ws[f'N{row_num}'] = trainee.total_workers or ''
            ws[f'O{row_num}'] = trainee.company_email or ''
            ws[f'P{row_num}'] = trainee.email or ''
            ws[f'Q{row_num}'] = trainee.phone_number or ''
            ws[f'R{row_num}'] = trainee.company_landline or ''
            ws[f'T{row_num}'] = mode_of_training
            ws[f'U{row_num}'] = f"#{trainee.schedule_id[-4:]}"
            
            # Add image if picture URL exists
            if trainee.picture_2x2_url:
                images_attempted += 1
                print(f"   📸 Processing image...")
                
                try:
                    if proxy_url:
                        img_url = f"{proxy_url}?url={trainee.picture_2x2_url}"
                        print(f"   🔗 Via proxy")
                    else:
                        img_url = trainee.picture_2x2_url
                        print(f"   🔗 Direct")
                    
                    image_data = fetch_image_with_retry(img_url)
                    
                    if image_data:
                        if len(image_data) < 100:
                            print(f"   ⚠️  Image too small")
                            images_failed += 1
                            failed_images.append(f"{trainee.first_name} {trainee.last_name}")
                            continue
                        
                        img = OpenpyxlImage(BytesIO(image_data))
                        img.width = 96
                        img.height = 96
                        ws.add_image(img, f'S{row_num}')
                        images_successful += 1
                        print(f"   ✅ Image added ({len(image_data)} bytes)")
                    else:
                        images_failed += 1
                        failed_images.append(f"{trainee.first_name} {trainee.last_name}")
                        print(f"   ❌ Failed to fetch image")
                        
                except Exception as e:
                    images_failed += 1
                    failed_images.append(f"{trainee.first_name} {trainee.last_name}")
                    print(f"   ❌ Error: {type(e).__name__}")
            else:
                print(f"   ⚪ No image URL")
        
        print(f"\n{'='*60}")
        print(f"📸 IMAGE PROCESSING SUMMARY")
        print(f"{'='*60}")
        print(f"Attempted: {images_attempted}")
        print(f"Successful: {images_successful}")
        print(f"Failed: {images_failed}")
        if failed_images:
            print(f"\nFailed images for:")
            for name in failed_images:
                print(f"  • {name}")
        print(f"{'='*60}\n")
        
        # Update master database
        print(f"{'='*60}")
        print(f"💾 UPDATING MASTER DATABASE")
        print(f"{'='*60}")
        
        append_to_master_database(
            course_name=course_name,
            training_dates=training_dates,
            schedule_id=schedule_id,
            trainee_count=trainee_count,
            male_count=male_count,
            female_count=female_count,
            mode_of_training=mode_of_training
        )
        
        # Copy all records from master to this export
        copy_master_database_to_export(wb)
        
        print(f"{'='*60}\n")
        
        # Save to BytesIO
        print(f"💾 Saving Excel file...")
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        wb.close()
        
        file_size = output.getbuffer().nbytes
        print(f"✅ Excel file generated ({file_size:,} bytes)")
        print(f"{'='*60}\n")
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=Originals{schedule_id[-4:]}.xlsx"
            }
        )
    
    except Exception as e:
        print(f"\n{'='*60}")
        print(f"❌ FATAL ERROR")
        print(f"{'='*60}")
        print(f"Error: {str(e)}")
        traceback.print_exc()
        print(f"{'='*60}\n")
        raise HTTPException(status_code=500, detail=str(e))

if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8000))
    print(f"🚀 Starting Excel Export Service with Master Database + Hostinger Backup on port {port}")
    uvicorn.run(app, host="0.0.0.0", port=port)




